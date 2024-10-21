# Versão em portugues
import streamlit as st
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import openpyxl
import os
from io import BytesIO  # Importar BytesIO
from matplotlib.backends.backend_pdf import PdfPages
from datetime import datetime


# Usar el backend 'Agg' para evitar problemas com Streamlit
import matplotlib
matplotlib.use('Agg')

# Lista de causas
causas = [
    'Causa 1', 'Causa 2', 'Causa 3', 'Causa 4', 'Causa 5', 
    'Causa 6', 'Causa 7', 'Causa 8', 'Causa 9', 'Causa 10', 
    'Causa 11', 'Causa 12', 'Causa 13', 'Causa 14', 'Causa 15',
    'Causa 16', 'Causa 17', 'Causa 18', 'Causa 19', 'Causa 20',
    'Causa 21', 'Causa 22', 'Causa 23', 'Causa 24', 'Causa 25'
]

# Função para generar Frequências aleatorias
def generar_ejemplo():
    # Generar un número de causas aleatorio entre 10 e 25
    num_causas = np.random.randint(10, 26)
    causas_seleccionadas = np.random.choice(causas, size=num_causas, replace=False)
    
    # Dividir as causas em 20% y 80%
    num_causas_20 = max(1, int(0.3 * num_causas))  # Ao menos uma causa estará en el 30%
    num_causas_80 = num_causas - num_causas_20
    
    # Asignar Frequências
    total_frecuencia = 600  # Frequência total (pode ser ajustada)
    
    # El 80% das Frequências para 20% das causas
    frecuencias_20 = np.random.randint(50, 100, size=num_causas_20)  # Frequências mais altas
    frecuencias_20 = frecuencias_20 / frecuencias_20.sum() * 0.7 * total_frecuencia  # Escalar 80% do total
    
    # O 20% das Frequências para 80% das causas
    frecuencias_80 = np.random.randint(1, 50, size=num_causas_80)  # Frecuencias más bajas
    frecuencias_80 = frecuencias_80 / frecuencias_80.sum() * 0.2 * total_frecuencia  # Escalar al 20% del total
    
    # Combinar as Frequências
    frecuencias = np.concatenate([frecuencias_20, frecuencias_80]).astype(int)
    
    # Criar DataFrame com as causas y Frequências
    df = pd.DataFrame({'Causa': causas_seleccionadas, 'Frecuencia': frecuencias})
    df = df.sort_values(by='Frecuencia', ascending=False)
    return df

# Configuração da navegação na barra lateral
st.sidebar.title("Navegação")
page = st.sidebar.radio("Ir a:", ["Inicio", "Aplicação", "Aprendendo", "Next" ])

# Control de Navegação con el estado de sesión
if "page" not in st.session_state:
    st.session_state.page = "Inicio"

# Navigation Manager
if page == "Inicio":
    st.session_state.page = "Inicio"
elif page == "Aplicação":
    st.session_state.page = "Aplicação"
elif page == "Aprendendo":
    st.session_state.page = "Aprendendo"
elif page == "Next":
    st.session_state.page = "Next"

# Mostrar conteudo segundo a página seleccionada
if st.session_state.page == "Inicio":
    # Página de inicio
    st.image("P4retoImage3.png", caption="Gerado con Adobe Firefly e Canva", use_column_width=True)  # Ruta da imagen local
    st.title("P4reto Chart 4.0")
    st.header('Instruções para preparar os dados de entrada para elaborar o Gráfico de Pareto')
    st.markdown("""
                1. Devemos carregar os dados em um arquivo de Excel (.xlsx ou .xls).
                2. Os cabeçalhos podem ser personalizados, mas devem seguir a seguinte ordem:
                - Primeira coluna: colocar as categorias (Falhas, demoras, ameaças, oportunidades, causas). Não existe
                    um limite de caracteres para esta categoria, mas um máximo de 25 caracteres mantém uma aparência legível.
                - Segunda coluna: colocar a frequência (tempo acumulado para cada evento).
                - O título do gráfico será o nome do arquivo sem extensão.
                - Os encabeçados das columnas serán utilizados como etiquetas dos eixos X (celula B1), Y (celula A1).
                3. Não deve haver células em branco ou valores nulos.
                4. O aplicativo se encarregará de ordenar a severidade dos eventos de acordo com a frequência em ordem decrescente.
                5. Um pequeno diferencial nesta proposta de interpretação de Pareto é a área
                sombreada correspondente a 80% das paradas, o que simplifica bastante a identificação
                dos aspectos aos quais prestar mais atenção, por isso a chamamos de "Pay Attention Zone".
                6. Gostaríamos muito de adiantar que essa simples diferenciação veio acompanhada de alguns
                insights muito valiosos, que estaremos desenvolvendo nos próximos dias. Como desafio, deixamos a seguinte pergunta: "Quão proativos ou reativos nos consideramos?"
                """)
    
#if st.button("Ir a la Aplicação"):
    #st.session_state.page = "Aplicação"

elif st.session_state.page == "Aplicação":
    # Página da Aplicação
    st.title("Painel de Análise de Falhas")
    st.write("Carga os dados de falhas e visualiza o gráfico.")
    
    # Botón para cargar dados
    uploaded_file = st.file_uploader("Cargar arquivo Excel", type=["xlsx"])

    # Inicializamos df_data como None
    df_data = None

    # Mostrar botão para gerar exemplo
    st.markdown("### Generar um exemplo aleatorio")
    st.info("Pode pulsar o botão para gerar os dados de falhas e visualizar um gráfico aleatorio.")
    # st.markdown("---")
    if st.button('Gerar exemplo'):
        df_data = generar_ejemplo()
        st.write("Exemplo Gerado com sucesso:")
        # Iniciar variaveis
        title_gp4 = 'Gráfico de Pareto 4.0'
        xlabel_causas = 'Causas'
        ylabel_frecuencia = 'Frequência'
    
    if uploaded_file is not None:
        # actulización de las variables para rotular el grafico
        wb = openpyxl.load_workbook(uploaded_file)
        ws = wb.active
        title_gp4 = os.path.splitext(uploaded_file.name)[0]

        # Obtener los valores de las celdas A1 y B1
        xlabel_causas = ws['A1'].value
        ylabel_frecuencia = ws['B1'].value

        # Se tem cargado o arquivo, leemos os dados num DataFrame
        df_data = pd.read_excel(uploaded_file)
        df_data.rename(columns={df_data.columns[0]: 'Causa', df_data.columns[1]: 'Frecuencia'}, inplace=True)
        st.write("Datos cargados com éxito:")
        #st.dataframe(df_data)

    if df_data is not None:      
        # Procesar dados para o gráfico
        try:
            if 'Frecuencia' in df_data.columns and 'Causa' in df_data.columns:
                df_data['Porcentaje'] = df_data['Frecuencia'] / df_data['Frecuencia'].sum() * 100
                df_data = df_data.sort_values(by='Frecuencia', ascending=False)
                frecuencia_max = df_data['Frecuencia'].max()
                df_data['Porcentaje Acumulado'] = df_data['Porcentaje'].cumsum()
                
                # Pay Attention
                condicion = 80
                for i, valor in enumerate(df_data['Porcentaje Acumulado']):
                    if valor >= condicion:
                        xmax = i + 0.5
                        break
                xmin = -1    
                
                # Criar o gráfico de Pareto
                fig, ax = plt.subplots(figsize=(10, 6))
                plt.title(title_gp4, fontsize=14, pad=10)
                ax.bar(df_data['Causa'], df_data['Frecuencia'], color='blue')
                ax.set_xlabel(xlabel_causas)
                ax.set_ylabel(ylabel_frecuencia)
                ax.set_xticklabels(df_data['Causa'], rotation=90, fontsize=8)
                ax2 = ax.twinx()
                ax2.plot(df_data['Causa'], df_data['Porcentaje Acumulado'], color='red', linestyle='-')
                ax2.set_ylabel("% Acumulado")
                ax.axvspan(xmin, xmax, color='gray', alpha=0.3)
                ax.text(x= xmax-0.3, y= frecuencia_max, s='Pay Attention\nZone', fontsize=12, color='white', horizontalalignment='right', verticalalignment='top', alpha=0.5)
                plt.tight_layout()
                plt.xlim(-1, len(df_data))
                plt.ylim(0, 105)
                         
                st.write("### Gráfico de Pareto 4.0")
                st.write("Eventos na área sombreada são responsaveis pelo 80% das paradas, Pay Attention!")
                st.pyplot(fig)

                # Descargar o gráfico como PDF
                pdf_buffer = BytesIO()
                with PdfPages(pdf_buffer) as pdf:
                    pdf.savefig(fig)
                    plt.close(fig)
            
                pdf_buffer.seek(0)
                st.download_button(
                    label="Descargar o Gráfico em PDF",
                    data=pdf_buffer,
                    file_name="pareto_chart.pdf",
                    mime="application/pdf"
                )

            else:
                st.warning("Certifique-se de que o arquivo contém as colunas corretas: 'Causa', 'Frecuencia'.")
        except KeyError as e:
            st.warning(f"Error ao processar os dados: {e}")
    else:
        # Si no se ha cargado el archivo, mostramos un mensaje
        st.write("### Aguardando os dados")
        st.info("Por favor, carga um arquivo Excel com os dados ou gerar um exemplo para visualizar o gráfico.")

if st.session_state.page == "Aprendendo":
    # Página Aprendendo utilizar Pareto
    st.image("P4retoImage3.png", caption="Gerado com Adobe Firefly e Canva", use_column_width=True)  # Ruta de la imagen local
    st.title("P4reto Chart 4.0")
    st.header('Instruções para aproveitar novos Insights do tradicional gráfico de Pareto')
    st.markdown("""
                - Pareto não é uma ferramenta nova, não pretendemos com este app retrabalhar séculos de conhecimento, já temos fontes suficientes de informação que inclusive nós mesmos consultamos na elaboração deste projeto. Mencionaremos dois links para quem estiver começando do absoluto zero, possa consultar e depois voltar aqui conosco.
                - Origem do princípio de Pareto: https://pt.wikipedia.org/wiki/Vilfredo_Pareto
                - Origem do gráfico de Pareto: https://pt.wikipedia.org/wiki/Diagrama_de_Pareto, o artigo está muito completo, mas tem um pequeno erro de transcrição. Mencionam sua criação no início da década de 90, quando na realidade o gráfico foi criado em 1937, pouco antes do início da década de 40.
                - Desde sua criação, os elementos do gráfico de Pareto sempre foram os mesmos: Eventos, Frequência e Porcentagem Acumulada. Posteriormente, realiza-se uma análise sobre os 20% de eventos que teoricamente causam 80% das falhas e, com base nessa análise, elaboram-se planos para corrigir a causa raiz dos desvios que geram esses eventos, que geram atraso ou prejudicam nossos processos. Até aqui, tudo bem (somos grandes crentes em práticas minimalistas), mas o que nos impedia de ter um pouco mais de ajuda visual? Respondemos rapidamente a essa pergunta com o motivo que nos levou a criar e compartilhar este app. Descobrimos que é uma mistura de limitações tecnológicas e conformismo criativo, já que no gráfico tradicional de Pareto é muito simples pegar um lápis e destacar quais são os eventos que cumprem a relação 80-20. Além disso, com as ferramentas tradicionais de plotagem, é um pouco complicado destacar automaticamente essa área que chamamos de "Pay Attention Zone". Com essa combinação de fatores, o gráfico de Pareto permaneceu inalterado por quase 90 anos.
                - "Pay Attention Zone" foi uma ideia inicialmente fugaz, mas que inesperadamente desencadeou uma avalanche de insights que iremos desenvolver progressivamente. Começaremos pelo mais básico, revisitando cada elemento tradicional e explicando detalhadamente a importância de identificar e agir rapidamente naqueles eventos que mais afetam nossos processos...
                - Convidamos você a nos acompanhar. 
                """)
    
    st.title("Como interpretar corretamente P4reto Chart 4.0")
    st.text('Estudemos paso a paso cada elemento')
    st.image("P4retoImage4.png", caption="Gerado com Adobe Firefly e Canva", use_column_width=True)  # Ruta de la imagen local
    st.header('Eventos')
    st.markdown("""
                **Eventos: Identificação e Registro para a Análise de Pareto**

                - **O que são os eventos?** No contexto do gráfico de Pareto, os "eventos" se referem a qualquer acontecimento, problema ou situação que interrompe ou afeta negativamente um processo. Esses eventos podem variar dependendo da indústria ou processo que estamos analisando, mas em geral, os eventos representam os incidentes que impactam a qualidade, a produtividade ou o desempenho operacional.

                - **Por que é importante registrar os eventos?** O sucesso de uma análise de Pareto depende da qualidade dos dados que alimentam o gráfico. Um registro correto e completo dos eventos permite visualizar de maneira precisa quais são as causas mais frequentes ou de maior impacto em um processo. Em outras palavras, identificar e registrar os eventos chave é o primeiro passo para realizar um diagnóstico eficaz que permita priorizar ações corretivas.

                - **Tipos de eventos a considerar:** Dependendo da indústria e do processo específico, diversos tipos de eventos podem ser registrados. A seguir, apresentamos alguns exemplos comuns na indústria de manufatura e na manutenção industrial:
                    - **Paradas de máquina:** Um evento crítico é qualquer tempo de inatividade não planejado devido a falhas técnicas.
                    - **Falhas de qualidade:** Estes são eventos onde o produto final não atende às especificações exigidas, o que gera desperdício ou retrabalho.
                    - **Atrasos na produção:** Se um processo produtivo é atrasado por problemas logísticos ou de abastecimento, é importante registrar esses eventos.
                    - **Acidentes ou incidentes de segurança:** Nas indústrias onde a segurança é prioritária, esses eventos são cruciais para identificar padrões e causas recorrentes.
                    - **Falhas de manutenção:** Incluem todas as vezes que um equipamento falha, seja por falhas mecânicas, elétricas ou outras, e que requerem intervenções de manutenção.

                - **Quais eventos NÃO devem ser registrados?** Frequentemente, comete-se o erro de registrar eventos que não contribuem para uma análise significativa do problema. Alguns exemplos de eventos que não devem ser considerados na análise de Pareto são:
                    - **Eventos isolados ou raros:** Seu impacto no sistema é mínimo e não geram um padrão que valha a pena ser abordado imediatamente.
                    - **Problemas menores ou irrelevantes:** Incidentes que não afetam de maneira significativa a qualidade ou a produtividade.
                    - **Eventos fora do controle do processo:** Se um evento não está diretamente relacionado com a operação interna do processo (por exemplo, uma falha externa de infraestrutura), estes podem distrair a análise do verdadeiro problema.

                **Registro de eventos: Boas práticas**

                Para garantir uma análise de Pareto eficaz, é crucial seguir boas práticas no registro de eventos:

                1. **Definir um critério claro:** Determinar que tipo de eventos devem ser registrados com base no impacto no processo (frequência, gravidade, tempo de inatividade, custo, etc.).
                2. **Manter consistência:** Os eventos devem ser registrados de maneira consistente, utilizando descrições claras e precisas.
                3. **Ferramentas de registro:** Utilizar software de gestão de manutenção ou qualidade que permita registrar eventos de forma rápida e eficiente.
                4. **Treinar o pessoal:** É essencial que todo o pessoal envolvido no processo de registro entenda a importância e o critério dos eventos que devem ser registrados.

                Essas são recomendações gerais. Estamos cientes de que cada processo é diferente, e cada um deve se adaptar às suas circunstâncias e requisitos legais específicos. O bom senso e o critério correto devem sempre prevalecer para que a solução não seja pior do que o problema que estamos tentando resolver.
                """)
    
    st.image("P4retoImage5.png", caption="Gerado com Adobe Firefly e Canva", use_column_width=True)  # Ruta de la imagen local
    st.header('Frequência')
    st.markdown("""
                **Frequência: Como Medir a Repetição dos Eventos**

                **O que é a frequência na análise de Pareto?**  
                A frequência se refere ao número de vezes que um evento ocorre em um período de tempo determinado. É um dos parâmetros chave no gráfico de Pareto, pois nos permite identificar quais são os eventos que acontecem com mais regularidade e, portanto, quais devem receber maior atenção ao implementar ações corretivas.

                **Importância da Frequência na Indústria:**  
                No ambiente industrial, medir a frequência de eventos como falhas de equipamentos, defeitos de produção ou interrupções de processos é fundamental. Os eventos mais frequentes costumam ter o maior impacto na produtividade ou nos custos operacionais. Por exemplo, se uma mesma falha em uma máquina ocorre várias vezes em um mês, esse evento provavelmente terá uma maior prioridade para ser corrigido do que um evento menos frequente, mas de alta gravidade.

                **Como registrar a frequência:**  
                É essencial ter um sistema confiável para medir quantas vezes um evento ocorre. Aqui estão algumas dicas para realizar esse acompanhamento:

                1. **Definir claramente os eventos:** Como mencionamos na seção anterior, os eventos devem ser bem definidos para evitar confusões. Por exemplo, se estamos medindo falhas de equipamentos, é necessário estabelecer critérios claros sobre o que constitui uma falha que deve ser registrada.
                
                2. **Estabelecer um período de tempo:** A frequência sempre deve ser medida dentro de um período específico, seja diário, semanal ou mensal, dependendo do processo que está sendo analisado. Isso permite comparações mais precisas e a detecção de tendências ao longo do tempo.

                3. **Usar software de registro:** Sistemas automatizados ou semi-automatizados, como os CMMS (sistemas de gestão de manutenção computadorizados), podem facilitar o acompanhamento da frequência dos eventos, garantindo que sejam registrados de maneira rápida e precisa.

                4. **Garantir que todos os eventos relevantes sejam registrados:** A omissão de um evento pode distorcer a análise e dificultar a identificação dos problemas mais frequentes.

                **Exemplos de Frequência na Indústria:**

                - **Falhas mecânicas:** Se um componente de uma máquina falha repetidamente durante uma semana, sua alta frequência será destacada no gráfico de Pareto, indicando que essa peça precisa de atenção prioritária.
                - **Defeitos na produção:** Se durante um processo produtivo falhas frequentes forem detectadas em uma linha específica, esses defeitos recorrentes se destacarão na análise de Pareto.
                - **Acidentes de trabalho:** Em um ambiente de segurança, registrar a frequência de incidentes pode ajudar a identificar áreas perigosas ou práticas inseguras recorrentes.

                **Como usar a frequência no Gráfico de Pareto:**  
                O gráfico de Pareto visualiza a frequência dos eventos por meio de barras ordenadas da maior para a menor. Os eventos que ocorrem com mais frequência terão as barras mais altas e serão os primeiros a serem considerados para ações corretivas. Além disso, a "frequência acumulada" pode ser visualizada através da curva de Pareto, que ajuda a identificar que porcentagem dos eventos (geralmente 80%) provém de um pequeno número de causas (geralmente 20%).

                **Foco na Ação:**  
                O ponto chave é que, ao identificar os eventos mais frequentes, podemos agir sobre eles com maior rapidez, permitindo melhorar a produtividade ou evitar perdas significativas. Frequentemente, os eventos mais frequentes representam as "vitórias rápidas" em termos de melhoria de processos.

                **Frequência vs. Impacto (ou Gravidade)**

                - **Frequência:** Como mencionamos, refere-se ao número de vezes que um evento ocorre em um período determinado. No entanto, um evento que ocorre com muita frequência, mas que tem um impacto baixo, pode não ser tão prioritário quanto um que ocorre com menos frequência, mas tem um impacto ou gravidade muito maior.

                - **Impacto ou Gravidade:** Neste contexto, o impacto está relacionado principalmente ao tempo de inatividade ou à magnitude das perdas causadas por um evento. Por exemplo, uma falha que interrompe a produção por 30 minutos terá um impacto maior do que uma que causa uma interrupção de apenas 5 minutos, mesmo que esta última ocorra mais vezes.

                **O dilema: O que priorizar?**

                - **Alta frequência, baixa gravidade:** São eventos que ocorrem regularmente, mas têm um impacto relativamente baixo. Um exemplo clássico na manutenção é uma falha de baixa criticidade que não interrompe o processo produtivo, mas gera pequenas interrupções.

                - **Baixa frequência, alta gravidade:** Esses eventos são menos frequentes, mas quando ocorrem, geram um impacto significativo, como um acidente grave ou uma parada total de produção que acarreta altos custos ou riscos de segurança.

                **Matriz Frequência-Gravidade:**  
                Uma estratégia útil é combinar ambos os aspectos em uma matriz, onde se cruzam a frequência e a gravidade. Isso permite visualizar de forma clara quais eventos devem ser priorizados para intervenção:

                - **Alta frequência, alta gravidade:** Prioridade máxima. Eventos que ocorrem com frequência e têm um impacto significativo, como paradas frequentes de máquinas críticas.
                - **Alta frequência, baixa gravidade:** Esses eventos podem ser incômodos, mas podem ser resolvidos com ações corretivas menos urgentes. No entanto, se se repetirem constantemente, podem acumular um impacto significativo ao longo do tempo.
                - **Baixa frequência, alta gravidade:** Embora esses eventos sejam raros, devem ter alta prioridade, pois as consequências de não tratá-los a tempo podem ser graves.
                - **Baixa frequência, baixa gravidade:** Esses são os eventos de menor prioridade, já que não afetam significativamente o desempenho ou a segurança.

                **Reflexão final**  
                Muitas vezes, teremos que lidar com a dualidade frequência-gravidade em nossas análises de Pareto. Por isso, é muito importante pensar além da frequência bruta e analisar o impacto dos eventos. Uma abordagem poderia ser aprender a calcular o impacto acumulado e como isso pode influenciar nas decisões de manutenção e produção.
                """)
    
    st.image("P4retoImage6.png", caption="Gerado com Adobe Firefly e Canva", use_column_width=True)  # Ruta de la imagen local
    st.header('Atenção')
    st.markdown("""
                **Pay Attention Zone**

                A incorporação da "Pay Attention Zone" representa uma inovação chave que está transformando nossa forma de abordar os dados. Tradicionalmente, o gráfico de Pareto se limitava a mostrar a relação 80/20, ordenando os eventos que geram a maioria dos problemas de forma decrescente. Com a ajuda da linha de Percentual Acumulado, podíamos deduzir empiricamente a interseção entre Percentuais e Eventos. No entanto, ao automatizar a geração de uma área sombreada que destaca visualmente os eventos responsáveis por 80% das paradas, otimizamos esse processo. O que começou como uma ideia aparentemente passageira se revelou um catalisador para novos insights, oferecendo uma visão muito mais profunda das causas subjacentes e de como resolvê-las.

                :rotating_light: Por favor, observe novamente a imagem que encabeça esta seção. "**Atenção**" é a fase da análise que separa o passado do futuro. Embora cada elemento da análise seja importante, "Atenção" é especialmente relevante porque concentra ambos os esforços: o estudo dos eventos, da frequência e da gravidade (passado) e o descobrimento das causas e a execução das soluções (futuro). Nesta fase, a intervenção de um especialista é indispensável. Observamos que a omissão da participação de um especialista afeta gravemente a eficácia das análises de causa raiz (**RCA**). A teoria é fundamental para fornecer estrutura e método ao processo, mas sem uma compreensão operativa detalhada, é fácil cair em soluções superficiais que não atacam a verdadeira raiz do problema.

                A **"Pay Attention Zone"** que implementamos tem o potencial de destacar visualmente os pontos críticos de forma direta, mas se as pessoas que realizam a análise carecem de conhecimento prático e experiência, corremos o risco de interpretar mal as informações ou de não explorar suficientemente as causas subjacentes.

                Como analogia, Sun Tzu dizia: **"Tática sem estratégia é o ruído antes da derrota."** Neste contexto, poderíamos dizer: **"Uma análise de dados sem conhecimento prático é um exercício vazio, condenado ao fracasso."** Isso destaca como o gráfico de Pareto e a "Pay Attention Zone" podem se tornar ferramentas poderosas, mas apenas se forem usadas corretamente por aqueles que compreendem tanto a teoria quanto a prática.:rotating_light:

                Esta "Pay Attention Zone" não só facilita a interpretação do gráfico, como também está mudando paradigmas ao permitir uma identificação mais intuitiva e rápida dos eventos críticos. Isso, por sua vez, impactou as ações que tomamos para resolver a causa raiz dos problemas, já que agora é mais fácil priorizar intervenções e recursos. Ao colocar maior ênfase nos eventos dentro desta zona, não apenas podemos otimizar a análise tradicional de Pareto, mas também desenvolver novas abordagens que se alinhem melhor à realidade das operações e da produção.

                Em resumo, essa simples mudança abriu um leque de oportunidades para reestruturar a análise, ajudando a tomar decisões mais estratégicas e baseadas em dados, oferecendo uma ferramenta ainda mais poderosa para a melhoria contínua no contexto industrial. Sabemos que pode parecer exagerado afirmar que uma simples área destacada pode mudar paradigmas, então explicaremos melhor essa afirmação na seção "Next" deste aplicativo.
                """)
    
    st.image("P4retoImage7.png", caption="Gerado com Adobe Firefly e Canva", use_column_width=True)  # Ruta de la imagen local
    st.header('Insights')
    st.markdown("""
                Uma vez que coletamos evidências suficientes por meio da análise de eventos e frequência, a próxima fase crítica é a descoberta de Insights. Os Insights são as descobertas chave que nos permitem entender as causas subjacentes por trás de cada evento. Para chegar a esse ponto, é fundamental que as etapas anteriores tenham sido executadas com precisão, garantindo que os dados sejam confiáveis e que haja um analista especialista capaz de realizar a fase de atenção com o rigor necessário.

                Consideramos importante destacar que o Pareto nos ajudará a encontrar os eventos mais relevantes ou prioritários, aqueles que precisam de "atenção" imediata, mas ele não é uma fonte de soluções. Para isso, uma vez detectadas as oportunidades, deve ser implementada uma estratégia de Análise de Causa Raiz (**RCA**), que contempla diferentes metodologias conforme as circunstâncias e as características de cada processo em particular. A mais apropriada deve ser selecionada, e é dessa análise de **RCA** que virão nossos tão desejados **Insights**.

                O Insight surge como resultado de uma análise profunda e estruturada, mas não se deve avançar sem antes validar as hipóteses. Aqui, é crucial verificar a veracidade de cada hipótese formulada sobre as causas dos eventos. Além disso, deve-se avaliar a viabilidade das possíveis soluções, garantindo que sejam práticas e eficazes no contexto industrial em que serão aplicadas. Ignorar essa etapa pode levar à implementação de ações incorretas ou ineficazes, perpetuando os problemas em vez de resolvê-los.

                O processo de descobrir um Insight não é apenas técnico, mas também estratégico, pois estabelece a ponte entre os dados coletados e as ações que serão tomadas para melhorar o sistema. Garantir que os Insights sejam precisos, baseados em dados verdadeiros e não em simples suposições, é fundamental para o sucesso na próxima fase da análise de Pareto.
                """)
    
    st.image("P4retoImage8.png", caption="Gerado com Adobe Firefly e Canva", use_column_width=True)  # Ruta de la imagen local
    st.header('Ação')
    st.markdown("""
                O Plano de Ação representa o ponto culminante de toda a análise anterior. É aqui que convergem os eventos registrados, a frequência com que ocorrem, a atenção prestada e os insights descobertos. O sucesso deste plano depende diretamente da precisão com que se tenha diagnosticado a causa raiz dos eventos, o que significa que quanto mais minucioso e exato tiver sido o processo, melhores serão os resultados obtidos nesta fase.

                O Plano de Ação não é apenas uma lista de tarefas; é uma guia estruturada que responde diretamente aos problemas identificados. Para ser eficaz, ele deve estar fundamentado em dados sólidos, uma avaliação objetiva e a verificação minuciosa das hipóteses levantadas nas fases anteriores. Cada ação deve estar alinhada com a realidade operacional e adaptada às capacidades da organização, com foco em resolver definitivamente as causas que originaram os eventos.

                A precisão no diagnóstico das causas subjacentes garante que o Plano de Ação esteja voltado para resolver os problemas na raiz, evitando recorrências e otimizando recursos. Portanto, cada etapa implementada deve estar bem justificada, com um foco preventivo, corretivo ou de melhoria, dependendo do tipo de evento e de seu impacto. A priorização é fundamental aqui, concentrando esforços nos eventos críticos identificados na "Pay Attention Zone" para maximizar os resultados e alcançar uma melhoria contínua.
                """)

# Mostrar el contenido según la página seleccionada
if st.session_state.page == "Next":
    # Página de inicio
    st.title("P4reto Chart 4.0")
    st.header('Pequenas mudanças poden gerar grandes aprendizados!')
    st.markdown("""
                **Conscientes da aparente irrelevância que uma simples área sombreada, nossa "Pay Attention Zone", poderia ter, apelamos à sua curiosidade e o convidamos a descobrir quanto valor esse pequeno detalhe pode agregar e como ele pode contribuir para enriquecer as informações que tradicionalmente obtemos de uma ferramenta tão clássica como o gráfico de Pareto.**

                - Desde sua criação, o gráfico de Pareto tem sido uma ferramenta visual para destacar os eventos que causam 80% de nossas paradas. Hoje, graças aos avanços tecnológicos, podemos destacar automaticamente essa área (nossa "Pay Attention Zone"), simplesmente adicionando cinco linhas de código ao popular script Python que gera esse tipo de gráfico. Para nossa surpresa, o que começou como um ajuste simples rapidamente nos revelou novos insights, abrindo portas para oportunidades mais amplas.
                - Relacionando essa pequena contribuição com outros estudos, como a curva de falhas potencial/funcional, começamos a identificar novas oportunidades, que iremos compartilhar progressivamente nesta aplicação, como se fosse um blog teórico-prático.
                - Sejam todos bem-vindos a esta aventura! Seus comentários, críticas e contribuições serão altamente valorizados. Esperamos que **P4reto** seja tão útil para vocês quanto é para nós. Estamos à disposição em: *elartedelmantenimientosuntzu@gmail.com*
                """)

# Obtener el día actual
hoy = datetime.now().day

# Mostrar el mensaje solo entre los días 29-05 de cada mes
if (29 <= hoy <= 31) or (1 <= hoy <= 5):
    st.markdown("""
        <hr>
        <p style='text-align: center; font-size: 12px;'>
            Se você aprecia nosso trabalho, considere fazer uma doação via PayPal. - elartedelmantenimientosuntzu@gmail.com<br>
            "Obrigado por contribuir com nosso projeto"
        </p>
        """, unsafe_allow_html=True)


# streamlit run Pareto-pt.py